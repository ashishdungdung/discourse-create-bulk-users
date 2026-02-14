import argparse
import logging
import os
import secrets
import string
from dataclasses import dataclass
from typing import Dict, Optional, Tuple

import openpyxl
import requests

LOG_FORMAT = "%(levelname)s: %(message)s"

REQUIRED_COLUMNS = ["username", "email", "name"]
OPTIONAL_COLUMNS = ["password"]
STATUS_COLUMNS = {
    "User Status": "user status",
    "API Response": "api response",
    "User ID": "user id",
    "Notes": "notes",
}


@dataclass
class Config:
    site_url: str
    api_key: str
    api_username: str
    timeout_seconds: int
    active: bool
    approved: bool
    suppress_welcome_message: bool
    dry_run: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Bulk-create Discourse users from an Excel file.")
    parser.add_argument("--file", default="users.xlsx", help="Path to the Excel workbook.")
    parser.add_argument("--site-url", default=os.getenv("DISCOURSE_SITE_URL", "").strip(), help="Discourse base URL.")
    parser.add_argument("--api-key", default=os.getenv("DISCOURSE_API_KEY", "").strip(), help="Discourse API key.")
    parser.add_argument(
        "--api-username",
        default=os.getenv("DISCOURSE_API_USERNAME", "").strip(),
        help="Admin username tied to the API key.",
    )
    parser.add_argument("--timeout", type=int, default=30, help="Request timeout in seconds.")
    parser.add_argument("--active", action="store_true", help="Create users as active.")
    parser.add_argument("--approved", action="store_true", help="Create users as approved.")
    parser.add_argument(
        "--suppress-welcome-message",
        action="store_true",
        help="Suppress Discourse welcome message on create.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Validate input and simulate API calls.")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    return parser.parse_args()


def build_config(args: argparse.Namespace) -> Config:
    site_url = args.site_url.rstrip("/")
    if not args.dry_run:
        missing = [
            name
            for name, value in [
                ("site-url", site_url),
                ("api-key", args.api_key),
                ("api-username", args.api_username),
            ]
            if not value
        ]
        if missing:
            raise ValueError(f"Missing required credentials/options: {', '.join(missing)}")

    return Config(
        site_url=site_url,
        api_key=args.api_key,
        api_username=args.api_username,
        timeout_seconds=args.timeout,
        active=args.active,
        approved=args.approved,
        suppress_welcome_message=args.suppress_welcome_message,
        dry_run=args.dry_run,
    )


def normalize(value: Optional[str]) -> str:
    return str(value or "").strip().lower()


def header_map(worksheet) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col_idx, cell in enumerate(worksheet[1], start=1):
        key = normalize(cell.value)
        if key:
            mapping[key] = col_idx
    return mapping


def ensure_columns(worksheet, mapping: Dict[str, int]) -> Dict[str, int]:
    missing_required = [col for col in REQUIRED_COLUMNS if col not in mapping]
    if missing_required:
        raise ValueError(f"Missing required columns in row 1: {', '.join(missing_required)}")

    max_col = worksheet.max_column
    for pretty_name, normalized in STATUS_COLUMNS.items():
        if normalized not in mapping:
            max_col += 1
            worksheet.cell(row=1, column=max_col, value=pretty_name)
            mapping[normalized] = max_col

    return mapping


def random_password(length: int = 20) -> str:
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def make_session(config: Config) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "Api-Key": config.api_key,
            "Api-Username": config.api_username,
            "Content-Type": "application/json",
        }
    )
    return session


def create_user(session: requests.Session, config: Config, user: Dict[str, str]) -> Tuple[bool, str, Optional[int]]:
    if config.dry_run:
        return True, "Dry run: request not sent", None

    payload = {
        "name": user["name"],
        "email": user["email"],
        "username": user["username"],
        "password": user["password"],
        "active": config.active,
        "approved": config.approved,
        "suppress_welcome_message": config.suppress_welcome_message,
    }

    url = f"{config.site_url}/users.json"
    try:
        response = session.post(url, json=payload, timeout=config.timeout_seconds)
        data = response.json() if response.content else {}
    except requests.RequestException as exc:
        return False, f"Request error: {exc}", None
    except ValueError:
        data = {"error": response.text}

    if response.ok and data.get("success"):
        user_id = data.get("user_id")
        return True, "Created", user_id

    errors = data.get("errors") or data.get("error") or response.text
    if isinstance(errors, list):
        error_text = "; ".join(str(item) for item in errors)
    else:
        error_text = str(errors)
    return False, f"HTTP {response.status_code}: {error_text}", None


def cell_value(worksheet, row: int, col: int) -> str:
    return str(worksheet.cell(row=row, column=col).value or "").strip()


def process_workbook(path: str, config: Config) -> None:
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active

    mapping = header_map(worksheet)
    mapping = ensure_columns(worksheet, mapping)

    session = make_session(config)

    status_col = mapping[STATUS_COLUMNS["User Status"]]
    response_col = mapping[STATUS_COLUMNS["API Response"]]
    user_id_col = mapping[STATUS_COLUMNS["User ID"]]
    notes_col = mapping[STATUS_COLUMNS["Notes"]]

    created = 0
    failed = 0
    skipped = 0

    for row_idx in range(2, worksheet.max_row + 1):
        row_data = {key: cell_value(worksheet, row_idx, mapping[key]) for key in REQUIRED_COLUMNS if key in mapping}
        for key in OPTIONAL_COLUMNS:
            row_data[key] = cell_value(worksheet, row_idx, mapping[key]) if key in mapping else ""

        if not any(row_data.values()):
            continue

        existing_user_id = cell_value(worksheet, row_idx, user_id_col)
        if existing_user_id:
            worksheet.cell(row=row_idx, column=status_col, value="Skipped")
            worksheet.cell(row=row_idx, column=notes_col, value="Already has User ID")
            skipped += 1
            continue

        if any(not row_data[field] for field in REQUIRED_COLUMNS):
            worksheet.cell(row=row_idx, column=status_col, value="Invalid row")
            worksheet.cell(row=row_idx, column=response_col, value="Missing one or more required fields")
            failed += 1
            continue

        generated_password = False
        if not row_data["password"]:
            row_data["password"] = random_password()
            generated_password = True

        ok, message, created_user_id = create_user(session, config, row_data)
        worksheet.cell(row=row_idx, column=status_col, value="Created" if ok else "Failed")
        worksheet.cell(row=row_idx, column=response_col, value=message)
        worksheet.cell(row=row_idx, column=user_id_col, value=created_user_id if created_user_id is not None else "")

        notes = "Generated password" if generated_password else ""
        worksheet.cell(row=row_idx, column=notes_col, value=notes)

        if ok:
            created += 1
        else:
            failed += 1

    workbook.save(path)
    logging.info("Completed. Created=%s Failed=%s Skipped=%s", created, failed, skipped)


def main() -> int:
    args = parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level), format=LOG_FORMAT)

    try:
        config = build_config(args)
        process_workbook(args.file, config)
    except Exception as exc:
        logging.error("%s", exc)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
